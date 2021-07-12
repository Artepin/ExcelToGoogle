import openpyxl
#import pyexcel
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font



class Exlib:

    fileread = openpyxl.load_workbook
    sheetid = 0

    def redFile(self, filePath):
        Exlib.fileread = openpyxl.load_workbook(filePath)

    def sheetID(self,id):
        Exlib.sheetid = id

    # получаем максимальное количество строк
    def getRows(self):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        return(sheet.max_row)

    # то же и с колоннами
    def getColumns(self):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        return(sheet.max_column)

    # читаем данные ячейки
    def getNumber(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        number = cell.value
        return(number)

    # узнаем букву колоны
    def columnLetter(self, columnnum):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        column = get_column_letter(columnnum)
        return (column)

    # получаем список всех объедененных зон
    def getMerged(self):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        mergedlist = sheet.merged_cells.ranges
        return mergedlist

    # эти функции возвращают ширину и высоту ячейки соответсвенно
    def getWidth(self, column):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        width = sheet.column_dimensions[column].width
        return width*7.1

    def getHeight(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        try:
            sheet.row_dimensions[cell1].height*1.34
        except TypeError:
            height = 21
        else:
            height = sheet.row_dimensions[cell1].height*1.34
        return height

    ################<<STYLES>>##################

    # тут стили текста
    # получаем данные о шрифте: Arial, Calibri и т. д.
    def getFont(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        font = cell.font.name
        return font

    # размер шрифта
    def getFontSize(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        fontsize = cell.font.size
        return int(fontsize)

    # его цвет
    def getFontColor(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        if len(str(cell.font.color.rgb))==8:
            fcolor = str(cell.font.color.rgb)
            ret = (fcolor[2:8])
        else:
            return False
        return ret

    # функции далее дают логичиский ответ да/нет, думаю по их названию можно понять,
    # проверкой чего они являются
    def getBold(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        boldstatus = cell.font.bold
        return bool(boldstatus)

    def getItalic(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        italicstatus = cell.font.italic
        return bool(italicstatus)

    def getStrikethrough(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        ststatus = cell.font.strikethrough
        return bool(ststatus)

    def getUnderline(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        undrlstatus = bool(cell.font.underline)
        if undrlstatus != False:
            return True

    # с текстом и булам покончили, дальше цвет фона ячейки
    # его можно брать отдельно по каналам
    def bgColorRed(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        val = cell.fill.start_color.index
        redvalue = (int(val[2:2 + 2], 16))
        return (redvalue)

    def bgColorGreen(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        val = cell.fill.start_color.index
        greenvalue = (int(val[4:4 + 2], 16))
        return (greenvalue)

    def bgColorBlue(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        val = cell.fill.start_color.index
        bluevalue = (int(val[6:6 + 2], 16))
        return (bluevalue)

    # или получить весь цвет ячейки
    def bgColor(self, cell1):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        #val = str(cell.fill.start_color.rgb)
        if len(str(cell.fill.start_color.rgb))==8:
            val = str(cell.fill.start_color.rgb)
            ret = (val[2:8])
        else:
            return False
        return ret

    ################<<BORDERS>>##################

    # получаем данные о границе
    def getBorder(self, cell1, border_orientation):
        file = Exlib.fileread
        sheet_id = Exlib.sheetid
        sheet = file.worksheets[sheet_id]
        cell = sheet[cell1]
        
        if border_orientation == "top":
            try:
                cell.border.top.color.rgb
            except AttributeError:
                clr = {}
            else:
                if len(str(cell.border.top.color.rgb))==8:
                    val = str(cell.border.top.color.rgb)
                    col = (val[2:8])
                    if col == "000000":
                        clr = {"red": 1, "green": 1, "blue": 1}
                    else:
                        clr = {"red": int(col[0:2], 16) / 255.0, "green": int(col[2:4], 16) / 255.0,
                            "blue": int(col[4:6], 16) / 255.0}
                else:
                    clr = {}
            border_unit = {
                'thin': "SOLID",
                'medium': "SOLID_MEDIUM",
                'thick': "SOLID_THICK",
                'dashed': "DASHED",
                'dotted': "DOTTED",
                'double': "DOUBLE",
                'dashDotDot': "DOTTED",
                'mediumDashDotDot': "DOTTED",
                'slantDashDot': "dashed",
                'mediumDashDot': "dashed",
                'mediumDashed': "dashed",
                'dashDot': "dashed",
                'hair': "DOTTED",
                'None': "NONE"
            }
            try:
                cell.border.top.border_style
            except AttributeError:
                st = 'NONE'
            else:
                st = border_unit[str(cell.border.top.border_style)]

        if border_orientation == "right":
            try:
                cell.border.right.color.rgb
            except AttributeError:
                clr = {}
            else:
                if len(str(cell.border.right.color.rgb))==8:
                    val = str(cell.border.right.color.rgb)
                    col = (val[2:8])
                    if col == "000000":
                        clr = {"red": 1, "green": 1, "blue": 1}
                    else:
                        clr = {"red": int(col[0:2], 16) / 255.0, "green": int(col[2:4], 16) / 255.0,
                            "blue": int(col[4:6], 16) / 255.0}
                else:
                    clr = {}
            border_unit = {
                'thin': "SOLID",
                'medium': "SOLID_MEDIUM",
                'thick': "SOLID_THICK",
                'dashed': "DASHED",
                'dotted': "DOTTED",
                'double': "DOUBLE",
                'dashDotDot': "DOTTED",
                'mediumDashDotDot': "DOTTED",
                'slantDashDot': "dashed",
                'mediumDashDot': "dashed",
                'mediumDashed': "dashed",
                'dashDot': "dashed",
                'hair': "DOTTED",
                'None': "NONE"
            }
            try:
                cell.border.right.border_style
            except AttributeError:
                st = 'NONE'
            else:
                st = border_unit[str(cell.border.right.border_style)]


        if border_orientation == "bottom":
            try:
                cell.border.bottom.color.rgb
            except AttributeError:
                clr = {}
            else:
                if len(str(cell.border.bottom.color.rgb))==8:
                    val = str(cell.border.bottom.color.rgb)
                    col = (val[2:8])
                    if col == "000000":
                        clr = {"red": 1, "green": 1, "blue": 1}
                    else:
                        clr = {"red": int(col[0:2], 16) / 255.0, "green": int(col[2:4], 16) / 255.0,
                            "blue": int(col[4:6], 16) / 255.0}
                else:
                    clr = {}
            border_unit = {
                'thin': "SOLID",
                'medium': "SOLID_MEDIUM",
                'thick': "SOLID_THICK",
                'dashed': "DASHED",
                'dotted': "DOTTED",
                'double': "DOUBLE",
                'dashDotDot': "DOTTED",
                'mediumDashDotDot': "DOTTED",
                'slantDashDot': "dashed",
                'mediumDashDot': "dashed",
                'mediumDashed': "dashed",
                'dashDot': "dashed",
                'hair': "DOTTED",
                'None': "NONE"
            }
            try:
                cell.border.bottom.border_style
            except AttributeError:
                st = 'NONE'
            else:
                st = border_unit[str(cell.border.bottom.border_style)]

        if border_orientation == "left":
            try:
                cell.border.left.color.rgb
            except AttributeError:
                clr = {}
            else:
                if len(str(cell.border.left.color.rgb)) == 8:
                    val = str(cell.border.left.color.rgb)
                    col = (val[2:8])
                    if col == "000000":
                        clr = {"red": 1, "green": 1, "blue": 1}
                    else:
                        clr = {"red": int(col[0:2], 16) / 255.0, "green": int(col[2:4], 16) / 255.0,
                               "blue": int(col[4:6], 16) / 255.0}
                else:
                    clr = {}
            border_unit = {
                'thin': "SOLID",
                'medium': "SOLID_MEDIUM",
                'thick': "SOLID_THICK",
                'dashed': "DASHED",
                'dotted': "DOTTED",
                'double': "DOUBLE",
                'dashDotDot': "DOTTED",
                'mediumDashDotDot': "DOTTED",
                'slantDashDot': "dashed",
                'mediumDashDot': "dashed",
                'mediumDashed': "dashed",
                'dashDot': "dashed",
                'hair': "DOTTED",
                'None': "NONE"
            }
            try:
                cell.border.left.border_style
            except AttributeError:
                st = 'NONE'
            else:
                st = border_unit[str(cell.border.left.border_style)]

        return {'style': st, 'width': 1, 'color': clr}



el = Exlib()